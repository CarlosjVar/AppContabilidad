from Clases import *
print("a")
from openpyxl import load_workbook
workbook=load_workbook(".\SpreadSheets\Contabilidad.xlsx")
print(workbook.sheetnames)

for pagina in self.workbook.sheetnames:
    self.ComboMes.addItem(pagina)
self.ComboMes.currentTextChanged.connect(self.cambioMes)
def CargaMeses(self, mesList):
    self.ComboMes.clear()
    for mes in mesList:
        self.ComboMes.addItem(mes)
 def cambioMes(self):
        print("hola")
        self.ValorIngresos.setText(str(self.sumaCeldas("J")))
        self.ValorNetos.setText((self.sumaCeldasS("J","D")))
        self.ValoGastos.setText(str(self.sumaCeldas("D")))
        self.ValorGastosPagar.setText((self.sumaCeldasS("D","E")))
    def sumaCeldas(self,columna):
        sheet = self.workbook[self.ComboMes.currentText()]
        n=2
        total=0
        print("inicio")
        while sheet[columna+str(n)].value:
            print("a")
            total=total+int(sheet[columna+str(n)].value)
            n+=1
        print(total)
        return total
    def sumaCeldasS(self,columna,columna2):
        sheet = self.workbook[self.ComboMes.currentText()]
        n=2
        total=0
        total2=0
        print("inicio")
        while sheet[columna+str(n)].value:
            total=total+int(sheet[columna+str(n)].value)
            n+=1
        n=2
        while sheet[columna2+str(n)].value:
            total2=total2+int(sheet[columna2+str(n)].value)
            n+=1
        return str(total-total2)