# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'C:\Users\carlo\Desktop\Contabilidad Liya\VentanaPrincipal.ui'
#
# Created by: PyQt5 UI code generator 5.13.2
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook

class Ui_MainWindow(object):
    workbook = load_workbook(".\SpreadSheets\Contabilidad.xlsx")
    def setupUi(self, MainWindow):

        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(766, 607)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(0, 0, 771, 561))
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.label = QtWidgets.QLabel(self.tab)
        self.label.setGeometry(QtCore.QRect(20, 30, 71, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.tab)
        self.label_2.setGeometry(QtCore.QRect(20, 120, 111, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.tab)
        self.label_3.setGeometry(QtCore.QRect(20, 60, 81, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.tab)
        self.label_4.setGeometry(QtCore.QRect(20, 90, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.ValorIngresos = QtWidgets.QLabel(self.tab)
        self.ValorIngresos.setGeometry(QtCore.QRect(170, 30, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.ValorIngresos.setFont(font)
        self.ValorIngresos.setText("")
        self.ValorIngresos.setObjectName("ValorIngresos")
        self.ValoGastos = QtWidgets.QLabel(self.tab)
        self.ValoGastos.setGeometry(QtCore.QRect(170, 60, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.ValoGastos.setFont(font)
        self.ValoGastos.setText("")
        self.ValoGastos.setObjectName("ValoGastos")
        self.ValorNetos = QtWidgets.QLabel(self.tab)
        self.ValorNetos.setGeometry(QtCore.QRect(170, 120, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.ValorNetos.setFont(font)
        self.ValorNetos.setText("")
        self.ValorNetos.setObjectName("ValorNetos")
        self.ValorGastosPagar = QtWidgets.QLabel(self.tab)
        self.ValorGastosPagar.setGeometry(QtCore.QRect(170, 90, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.ValorGastosPagar.setFont(font)
        self.ValorGastosPagar.setText("")
        self.ValorGastosPagar.setObjectName("ValorGastosPagar")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.tabWidget.addTab(self.tab_2, "")
        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.tabWidget.addTab(self.tab_3, "")
        self.ComboMes = QtWidgets.QComboBox(self.centralwidget)
        self.ComboMes.setGeometry(QtCore.QRect(50, 570, 69, 22))
        self.ComboMes.setObjectName("ComboMes")
        for pagina in self.workbook.sheetnames:
            self.ComboMes.addItem(pagina)
        self.ComboMes.currentTextChanged.connect(self.cambioMes)
        self.LabelMes = QtWidgets.QLabel(self.centralwidget)
        self.LabelMes.setGeometry(QtCore.QRect(10, 560, 61, 31))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.LabelMes.setFont(font)
        self.LabelMes.setObjectName("LabelMes")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "Ingresos"))
        self.label_2.setText(_translate("MainWindow", "Ingresos Netos"))
        self.label_3.setText(_translate("MainWindow", "Gastos"))
        self.label_4.setText(_translate("MainWindow", "Gastos por pagar"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "Ventana Resumen"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("MainWindow", "Desglose Ingresos"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("MainWindow", "Desglose Egresos"))
        self.LabelMes.setText(_translate("MainWindow", "Mes"))
    def cambioMes(self):
        print("hola")
        self.ValorIngresos.setText(str(self.sumaCeldas("J")))
        self.ValorNetos.setText((self.sumaCeldasS("J","D")))
        self.ValoGastos.setText(str(self.sumaCeldas("D")))
        self.ValorGastosPagar.setText((self.sumaCeldasS("D","E")))
    def sumaCeldas(self,columna):
        sheet = self.workbook[self.ComboMes.currentText()]
        n=2
        total=0
        print("inicio")
        while sheet[columna+str(n)].value:
            print("a")
            total=total+int(sheet[columna+str(n)].value)
            n+=1
        print(total)
        return total
    def sumaCeldasS(self,columna,columna2):
        sheet = self.workbook[self.ComboMes.currentText()]
        n=2
        total=0
        total2=0
        print("inicio")
        while sheet[columna+str(n)].value:
            total=total+int(sheet[columna+str(n)].value)
            n+=1
        n=2
        while sheet[columna2+str(n)].value:
            total2=total2+int(sheet[columna2+str(n)].value)
            n+=1
        return str(total-total2)




if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
